from openpyxl import Workbook
import openpyxl
from pathlib import Path
from logging import Logger
from applib.with_mixin import WithMixin



class XlsxFile(WithMixin):
    COLUMNS = (
        'sku', 
        'item',
        'description', 
        'size', 
        'weight', 
        'rrp', 
        'discount_price', 
        'promotion_price',
    )


    def __init__(self, logger: Logger, filepath: str) -> None:
        self._filepath = filepath
        self._logger = logger
        self._prepare_file()
        try:
            self._workbook = openpyxl.load_workbook(self._filepath)
        except Exception:
            self._logger.error('Workbook %s not found after preparation.', self._filepath)
            raise
        self._sheet = self._workbook.active


    def add_row(self, data: tuple) -> None:
        """Add row to end of xlsx file."""
        if not self._sheet:
            self._logger.warning('No active sheet to add row.')
            return
        
        if len(data) != len(self.COLUMNS):
            self._logger.info(
                'Row length mismatch expected: %d, got: %d', 
                len(self.COLUMNS), 
                len(data)
            )
            return

        self._sheet.append(data)
        self._workbook.save(self._filepath)


    def rows(self) -> tuple:
        """Return rows from xlsx file."""
        if not self._sheet:
            self._logger.warning('no active sheet. Return an empty tuple.')
            return tuple()

        return tuple(self._sheet.iter_rows(values_only=True))


    def _prepare_file(self) -> None:
        """Prepare path for file."""
        path = Path(self._filepath)

        if Path.exists(path):
            self._logger.info('File %s already exists.', self._filepath)
            return
        
        self._logger.info('Creating file %s', self._filepath)
        path.parent.mkdir(parents=True, exist_ok=True)
        self._create_with_columns()


    def _create_with_columns(self) -> None:
        """Create a new xlsx file with header columns."""
        workbook = Workbook()
        sheet = workbook.active
        
        if sheet is not None:
            sheet.append(self.COLUMNS)
        
        workbook.save(self._filepath)


    def close(self) -> None:
        """Close workbook."""
        if self._workbook:
            self._logger.debug('Closing %s', self._filepath)
            self._workbook.close()
